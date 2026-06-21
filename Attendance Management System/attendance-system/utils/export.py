"""
utils/export.py
---------------
Export attendance records to Excel (.xlsx) and PDF.
"""

from __future__ import annotations
import os
from datetime import datetime


def export_to_excel(records: list[dict], subject_name: str, session_title: str, filepath: str) -> str:
    """Export attendance records to an Excel file. Returns the saved filepath."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance"

    # ── Styles ──────────────────────────────────────────────────────────────
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", fgColor="1A1A2E")
    title_font    = Font(bold=True, size=14, color="1A1A2E")
    sub_font      = Font(size=10, italic=True, color="555555")
    center_align  = Alignment(horizontal="center", vertical="center")
    thin_border   = Border(
        left   = Side(style="thin"),
        right  = Side(style="thin"),
        top    = Side(style="thin"),
        bottom = Side(style="thin")
    )
    alt_fill = PatternFill("solid", fgColor="F0F4FF")

    # ── Title block ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    ws["A1"] = f"Attendance Report — {subject_name}"
    ws["A1"].font = title_font
    ws["A1"].alignment = center_align

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Session: {session_title}  |  Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = sub_font
    ws["A2"].alignment = center_align

    ws.append([])  # spacer row

    # ── Column headers ───────────────────────────────────────────────────────
    headers = ["#", "Full Name", "Reg. No", "Program", "Method", "Signed At"]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border

    # ── Data rows ────────────────────────────────────────────────────────────
    for i, rec in enumerate(records, 1):
        row_data = [
            i,
            rec.get("full_name", ""),
            rec.get("reg_no", ""),
            rec.get("program", ""),
            rec.get("method", "manual").capitalize(),
            rec.get("signed_at", ""),
        ]
        ws.append(row_data)
        row_idx = 4 + i
        row_fill = alt_fill if i % 2 == 0 else None
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = center_align if col in (1, 5) else Alignment(vertical="center")
            if row_fill:
                cell.fill = row_fill

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = [5, 28, 18, 22, 12, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Summary row ──────────────────────────────────────────────────────────
    ws.append([])
    ws.append(["", f"Total Present: {len(records)}", "", "", "", ""])
    summary_cell = ws.cell(row=ws.max_row, column=2)
    summary_cell.font = Font(bold=True, color="1A1A2E")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[4].height = 22

    wb.save(filepath)
    return filepath


def export_to_pdf(records: list[dict], subject_name: str, session_title: str, filepath: str) -> str:
    """Export attendance records to a PDF file. Returns the saved filepath."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        )
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise RuntimeError("reportlab is not installed. Run: pip install reportlab")

    doc = SimpleDocTemplate(
        filepath,
        pagesize = landscape(A4),
        rightMargin = 20*mm,
        leftMargin  = 20*mm,
        topMargin   = 20*mm,
        bottomMargin= 20*mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title2", parent=styles["Title"],
        fontSize=16, textColor=colors.HexColor("#1A1A2E"),
        spaceAfter=6
    )
    sub_style = ParagraphStyle(
        "Sub", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#555555"),
        spaceAfter=12
    )

    elements = [
        Paragraph(f"Attendance Report — {subject_name}", title_style),
        Paragraph(
            f"Session: <b>{session_title}</b> &nbsp;&nbsp;|&nbsp;&nbsp; "
            f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            sub_style
        ),
        Spacer(1, 8)
    ]

    # ── Table data ───────────────────────────────────────────────────────────
    header = ["#", "Full Name", "Reg. No", "Program", "Method", "Signed At"]
    data   = [header]
    for i, rec in enumerate(records, 1):
        data.append([
            str(i),
            rec.get("full_name", ""),
            rec.get("reg_no", ""),
            rec.get("program", ""),
            rec.get("method", "manual").capitalize(),
            rec.get("signed_at", ""),
        ])

    table = Table(data, colWidths=[15*mm, 55*mm, 35*mm, 45*mm, 25*mm, 45*mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1A2E")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, 0), 10),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",     (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#F7F9FF"), colors.HexColor("#EAEFFF")]),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWHEIGHT",  (0, 0), (-1, -1), 20),
        ("LEFTPADDING",  (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))
    summary_style = ParagraphStyle(
        "Summary", parent=styles["Normal"],
        fontSize=10, textColor=colors.HexColor("#1A1A2E")
    )
    elements.append(Paragraph(f"<b>Total Present: {len(records)}</b>", summary_style))

    doc.build(elements)
    return filepath


def get_default_export_path(subject_code: str, session_title: str, ext: str) -> str:
    """Return a default export filepath in the user's home/Downloads folder."""
    safe_title = "".join(c for c in session_title if c.isalnum() or c in " -_").strip()
    filename   = f"{subject_code}_{safe_title}_{datetime.now().strftime('%Y%m%d_%H%M')}.{ext}"
    downloads  = os.path.join(os.path.expanduser("~"), "Downloads")
    os.makedirs(downloads, exist_ok=True)
    return os.path.join(downloads, filename)
